"""
=============================================================
 Sage Ventures – Multifamily Analytics
 PHASE 4: Automated Excel Report Generator
 Run:    python excel_report.py
 Needs:  pip install pandas openpyxl
 Run AFTER Phase 1, 2 & 3
=============================================================
 Produces a fully formatted, multi-sheet Excel workbook —
 exactly the kind of Monday morning report a DA would
 schedule and auto-email to leadership.
=============================================================
"""

import sqlite3
import pandas as pd
import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint

DB_PATH    = "data/sage_ventures.db"
OUTPUT_DIR = "outputs"
os.makedirs(OUTPUT_DIR, exist_ok=True)

TODAY     = date.today()
RPT_DATE  = TODAY.strftime("%B %d, %Y")
OUT_FILE  = f"{OUTPUT_DIR}/SageVentures_Analytics_Report_{TODAY.strftime('%Y%m%d')}.xlsx"

# ── BRAND COLORS ──────────────────────────────────────────────────────────────
DARK_GREEN  = "1A3A2A"
MID_GREEN   = "1A6B3A"
LIGHT_GREEN = "E8F5E9"
ACCENT_GOLD = "C8A951"
WHITE       = "FFFFFF"
LIGHT_GRAY  = "F5F5F5"
MID_GRAY    = "CCCCCC"
DARK_GRAY   = "444444"
RED_LIGHT   = "FFEBEE"
RED_DARK    = "C62828"
AMBER_LIGHT = "FFF8E1"
AMBER_DARK  = "E65100"

def side(style="thin", color="CCCCCC"):
    return Side(style=style, color=color)

def border(all_sides="thin"):
    s = side(all_sides)
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=DARK_GRAY, size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic,
                name="Calibri")

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_col_widths(ws, widths: dict):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

def header_row(ws, row, values, bg=DARK_GREEN, fg=WHITE, bold=True, size=11):
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill    = fill(bg)
        cell.font    = font(bold=bold, color=fg, size=size)
        cell.border  = border()
        cell.alignment = align("center")

def data_row(ws, row, values, bg=WHITE, fmt_map=None, bold=False):
    fmt_map = fmt_map or {}
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill      = fill(bg)
        cell.font      = font(bold=bold, color=DARK_GRAY)
        cell.border    = border()
        cell.alignment = align("left" if isinstance(val,str) else "right")
        if col in fmt_map:
            cell.number_format = fmt_map[col]

def title_block(ws, title, subtitle, logo_text="SAGE VENTURES"):
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = logo_text
    c.font      = Font(bold=True, size=16, color=WHITE, name="Calibri")
    c.fill      = fill(DARK_GREEN)
    c.alignment = align("left","center")

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value     = title
    c.font      = Font(bold=True, size=13, color=DARK_GREEN, name="Calibri")
    c.fill      = fill(LIGHT_GREEN)
    c.alignment = align("left","center")

    ws.merge_cells("A3:H3")
    c = ws["A3"]
    c.value     = f"Report Date: {RPT_DATE}  |  {subtitle}"
    c.font      = Font(size=10, color=DARK_GRAY, italic=True, name="Calibri")
    c.fill      = fill(LIGHT_GRAY)
    c.alignment = align("left","center")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 18


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — EXECUTIVE SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
def sheet_executive_summary(wb, conn):
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False
    title_block(ws, "Portfolio Executive Summary",
                "Multifamily Analytics Dashboard — All Properties")

    # KPI data from SQL
    kpi = pd.read_sql("""
    SELECT
      COUNT(DISTINCT p.property_id)                              AS total_properties,
      COUNT(DISTINCT u.unit_id)                                  AS total_units,
      COUNT(DISTINCT t.tenant_id)                                AS occupied_units,
      ROUND(COUNT(DISTINCT t.tenant_id)*100.0/COUNT(DISTINCT u.unit_id),1) AS occ_pct,
      ROUND(AVG(u.market_rent),0)                                AS avg_market_rent,
      ROUND(AVG(t.monthly_rent),0)                               AS avg_actual_rent,
      ROUND(SUM(t.monthly_rent),0)                               AS total_monthly_rev
    FROM properties p
    JOIN units u      ON p.property_id = u.property_id
    LEFT JOIN tenants t ON u.unit_id  = t.unit_id AND t.status='Active'
    """, conn).iloc[0]

    # KPI cards (row 5 onwards)
    ws.row_dimensions[4].height = 10
    kpi_labels = [
        ("Total Properties",   kpi["total_properties"],   "",          WHITE),
        ("Total Units",        kpi["total_units"],         "",          WHITE),
        ("Occupied Units",     kpi["occupied_units"],      "",          WHITE),
        ("Portfolio Occupancy",f"{kpi['occ_pct']}%",      "Target 93%",LIGHT_GREEN),
        ("Avg Market Rent",    f"${kpi['avg_market_rent']:,.0f}", "",   WHITE),
        ("Avg Actual Rent",    f"${kpi['avg_actual_rent']:,.0f}", "Loss-to-Lease",WHITE),
        ("Monthly Revenue",    f"${kpi['total_monthly_rev']:,.0f}","",  LIGHT_GREEN),
    ]

    for idx, (label, value, note, bg) in enumerate(kpi_labels):
        row = 5 + (idx // 4) * 4
        col = 1 + (idx % 4) * 2
        ws.merge_cells(start_row=row,   start_column=col, end_row=row,   end_column=col+1)
        ws.merge_cells(start_row=row+1, start_column=col, end_row=row+1, end_column=col+1)
        ws.merge_cells(start_row=row+2, start_column=col, end_row=row+2, end_column=col+1)

        lc = ws.cell(row=row,   column=col, value=label)
        vc = ws.cell(row=row+1, column=col, value=value)
        nc = ws.cell(row=row+2, column=col, value=note)

        lc.font = Font(size=9, color=DARK_GRAY, name="Calibri")
        vc.font = Font(bold=True, size=18, color=DARK_GREEN, name="Calibri")
        nc.font = Font(size=9, italic=True, color=DARK_GRAY, name="Calibri")
        for cell in [lc, vc, nc]:
            cell.fill = fill(bg)
            cell.alignment = align("center")

    # Property performance table
    start_r = 14
    ws.cell(row=start_r, column=1, value="Property Performance Summary").font = \
        Font(bold=True, size=11, color=DARK_GREEN, name="Calibri")

    cols = ["Property","City","Total Units","Occupied","Occ %",
            "Avg Market Rent","Avg Actual Rent","Monthly Revenue"]
    header_row(ws, start_r+1, cols)

    props = pd.read_sql("""
    SELECT p.name, p.city, p.total_units,
      COUNT(t.tenant_id) occupied,
      ROUND(COUNT(t.tenant_id)*100.0/p.total_units,1) occ_pct,
      ROUND(AVG(u.market_rent),0) avg_mkt,
      ROUND(AVG(t.monthly_rent),0) avg_actual,
      ROUND(SUM(t.monthly_rent),0) monthly_rev
    FROM properties p
    JOIN units u ON p.property_id=u.property_id
    LEFT JOIN tenants t ON u.unit_id=t.unit_id AND t.status='Active'
    GROUP BY p.property_id ORDER BY monthly_rev DESC
    """, conn)

    fmt = {5:"0.0%", 6:"$#,##0", 7:"$#,##0", 8:"$#,##0"}
    for i, (_, row) in enumerate(props.iterrows()):
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        occ = row["occ_pct"] / 100
        data_row(ws, start_r+2+i,
                 [row["name"],row["city"],row["total_units"],row["occupied"],
                  occ, row["avg_mkt"], row["avg_actual"], row["monthly_rev"]],
                 bg=bg, fmt_map=fmt)
        # Color-code occupancy
        occ_cell = ws.cell(row=start_r+2+i, column=5)
        if occ < 0.80:
            occ_cell.fill = fill(RED_LIGHT)
            occ_cell.font = font(color=RED_DARK, bold=True)
        elif occ >= 0.90:
            occ_cell.fill = fill(LIGHT_GREEN)
            occ_cell.font = font(color=MID_GREEN, bold=True)

    set_col_widths(ws, {"A":28,"B":14,"C":12,"D":10,"E":10,
                        "F":16,"G":16,"H":16})


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — RENT ROLL
# ══════════════════════════════════════════════════════════════════════════════
def sheet_rent_roll(wb, conn):
    ws = wb.create_sheet("Rent Roll")
    ws.sheet_view.showGridLines = False
    title_block(ws, "Rent Roll — Current Month",
                "All active tenants, lease terms and payment status")

    last_month = pd.read_sql(
        "SELECT MAX(period_month) mo FROM rent_roll", conn).iloc[0]["mo"]

    df = pd.read_sql(f"""
    SELECT
        p.name property, p.city,
        u.unit_number, u.unit_type,
        t.full_name tenant,
        t.lease_start, t.lease_end,
        t.monthly_rent,
        u.market_rent,
        ROUND(u.market_rent - t.monthly_rent,0) loss_to_lease,
        r.amount_due, r.amount_paid,
        ROUND(r.amount_due-r.amount_paid,0) balance,
        r.status payment_status
    FROM rent_roll r
    JOIN tenants t    ON r.tenant_id   = t.tenant_id
    JOIN units u      ON r.unit_id     = u.unit_id
    JOIN properties p ON r.property_id = p.property_id
    WHERE r.period_month = '{last_month}'
    ORDER BY p.name, u.unit_number
    """, conn)

    cols = ["Property","City","Unit","Type","Tenant","Lease Start","Lease End",
            "Monthly Rent","Market Rent","Loss-to-Lease",
            "Amount Due","Amount Paid","Balance","Payment Status"]
    header_row(ws, 5, cols)

    status_colors = {
        "Paid":       (LIGHT_GREEN, MID_GREEN),
        "Partial":    (AMBER_LIGHT, AMBER_DARK),
        "Delinquent": (RED_LIGHT,   RED_DARK),
    }
    money_cols = {8:"$#,##0",9:"$#,##0",10:"$#,##0",11:"$#,##0",12:"$#,##0",13:"$#,##0"}

    for i, (_, row) in enumerate(df.iterrows()):
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        data_row(ws, 6+i,
                 [row["property"],row["city"],row["unit_number"],row["unit_type"],
                  row["tenant"],row["lease_start"],row["lease_end"],
                  row["monthly_rent"],row["market_rent"],row["loss_to_lease"],
                  row["amount_due"],row["amount_paid"],row["balance"],
                  row["payment_status"]],
                 bg=bg, fmt_map=money_cols)
        # Color status cell
        status_cell = ws.cell(row=6+i, column=14)
        st = row["payment_status"]
        if st in status_colors:
            bg_c, fg_c = status_colors[st]
            status_cell.fill = fill(bg_c)
            status_cell.font = font(color=fg_c, bold=True)

    set_col_widths(ws, {"A":24,"B":14,"C":7,"D":12,"E":22,
                        "F":12,"G":12,"H":13,"I":13,"J":13,
                        "K":12,"L":12,"M":10,"N":13})
    ws.freeze_panes = "A6"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — LEASE EXPIRATION PIPELINE
# ══════════════════════════════════════════════════════════════════════════════
def sheet_lease_expiration(wb, conn):
    ws = wb.create_sheet("Lease Expiration")
    ws.sheet_view.showGridLines = False
    title_block(ws, "Lease Expiration Pipeline",
                "Renewal risk analysis — next 90 days")

    df = pd.read_sql("""
    SELECT
        p.name property, p.city,
        t.full_name tenant,
        u.unit_number, u.unit_type,
        t.monthly_rent,
        t.lease_end,
        CAST(julianday(t.lease_end)-julianday('now') AS INTEGER) days_to_expiry,
        CASE
          WHEN julianday(t.lease_end)-julianday('now') < 0    THEN 'Expired'
          WHEN julianday(t.lease_end)-julianday('now') <= 30  THEN '0-30 Days'
          WHEN julianday(t.lease_end)-julianday('now') <= 60  THEN '31-60 Days'
          ELSE '61-90 Days'
        END bucket
    FROM tenants t
    JOIN units u      ON t.unit_id     = u.unit_id
    JOIN properties p ON u.property_id = p.property_id
    WHERE julianday(t.lease_end)-julianday('now') <= 90
      AND t.status = 'Active'
    ORDER BY days_to_expiry
    """, conn)

    cols = ["Property","City","Tenant","Unit","Type",
            "Monthly Rent","Lease End","Days to Expiry","Bucket"]
    header_row(ws, 5, cols)

    bucket_colors = {
        "Expired":    (RED_LIGHT,   RED_DARK),
        "0-30 Days":  (RED_LIGHT,   RED_DARK),
        "31-60 Days": (AMBER_LIGHT, AMBER_DARK),
        "61-90 Days": (LIGHT_GREEN, MID_GREEN),
    }
    for i, (_, row) in enumerate(df.iterrows()):
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        data_row(ws, 6+i,
                 [row["property"],row["city"],row["tenant"],
                  row["unit_number"],row["unit_type"],row["monthly_rent"],
                  row["lease_end"],row["days_to_expiry"],row["bucket"]],
                 bg=bg, fmt_map={6:"$#,##0"})
        bc = ws.cell(row=6+i, column=9)
        bucket = row["bucket"]
        if bucket in bucket_colors:
            bg_c, fg_c = bucket_colors[bucket]
            bc.fill = fill(bg_c)
            bc.font = font(color=fg_c, bold=True)

    set_col_widths(ws, {"A":24,"B":14,"C":22,"D":7,"E":12,
                        "F":13,"G":12,"H":14,"I":12})
    ws.freeze_panes = "A6"


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — DELINQUENCY REPORT
# ══════════════════════════════════════════════════════════════════════════════
def sheet_delinquency(wb, conn):
    ws = wb.create_sheet("Delinquency Report")
    ws.sheet_view.showGridLines = False
    title_block(ws, "Delinquency Report — Current Month",
                "Accounts with unpaid or partial balances")

    last_month = pd.read_sql(
        "SELECT MAX(period_month) mo FROM rent_roll", conn).iloc[0]["mo"]

    df = pd.read_sql(f"""
    SELECT
        p.name property, p.city,
        t.full_name tenant,
        u.unit_number, u.unit_type,
        r.period_month,
        r.amount_due, r.amount_paid,
        ROUND(r.amount_due-r.amount_paid,0) balance_owed,
        r.status,
        t.lease_end
    FROM rent_roll r
    JOIN tenants t    ON r.tenant_id   = t.tenant_id
    JOIN units u      ON r.unit_id     = u.unit_id
    JOIN properties p ON r.property_id = p.property_id
    WHERE r.status IN ('Delinquent','Partial')
      AND r.period_month = '{last_month}'
    ORDER BY balance_owed DESC
    """, conn)

    # Summary box
    ws.cell(row=5, column=1, value="Total Delinquent Accounts:").font = font(bold=True)
    ws.cell(row=5, column=2, value=len(df[df["status"]=="Delinquent"])).font = font(color=RED_DARK, bold=True)
    ws.cell(row=5, column=3, value="Total Exposure:").font = font(bold=True)
    ws.cell(row=5, column=4, value=df["balance_owed"].sum()).font = font(color=RED_DARK, bold=True)
    ws.cell(row=5, column=4).number_format = "$#,##0"

    cols = ["Property","City","Tenant","Unit","Type",
            "Period","Amount Due","Amount Paid","Balance Owed","Status","Lease End"]
    header_row(ws, 7, cols, bg=RED_DARK)

    for i, (_, row) in enumerate(df.iterrows()):
        bg = RED_LIGHT if row["status"]=="Delinquent" else AMBER_LIGHT
        data_row(ws, 8+i,
                 [row["property"],row["city"],row["tenant"],
                  row["unit_number"],row["unit_type"],row["period_month"],
                  row["amount_due"],row["amount_paid"],row["balance_owed"],
                  row["status"],row["lease_end"]],
                 bg=bg, fmt_map={7:"$#,##0",8:"$#,##0",9:"$#,##0"})

    set_col_widths(ws, {"A":24,"B":14,"C":22,"D":7,"E":12,
                        "F":10,"G":12,"H":12,"I":12,"J":12,"K":12})


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — YSR REVENUE TREND (with chart)
# ══════════════════════════════════════════════════════════════════════════════
def sheet_ysr(wb, conn):
    ws = wb.create_sheet("YSR Revenue Trend")
    ws.sheet_view.showGridLines = False
    title_block(ws, "Yield Summary Report (YSR) — Monthly Revenue Trend",
                "Gross Potential Rent vs Collected Rent | 15-month view")

    df = pd.read_sql("""
    SELECT
        period_month,
        ROUND(SUM(amount_due),0)  gross_potential_rent,
        ROUND(SUM(amount_paid),0) collected_rent,
        ROUND(SUM(amount_due)-SUM(amount_paid),0) uncollected,
        ROUND(SUM(amount_paid)*100.0/SUM(amount_due),1) collection_rate_pct
    FROM rent_roll
    GROUP BY period_month
    ORDER BY period_month
    """, conn)

    cols = ["Period","Gross Potential Rent","Collected Rent",
            "Uncollected","Collection Rate %"]
    header_row(ws, 5, cols)

    for i, (_, row) in enumerate(df.iterrows()):
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        data_row(ws, 6+i,
                 [row["period_month"],row["gross_potential_rent"],
                  row["collected_rent"],row["uncollected"],
                  row["collection_rate_pct"]/100],
                 bg=bg, fmt_map={2:"$#,##0",3:"$#,##0",4:"$#,##0",5:"0.0%"})

    # Bar Chart: GPR vs Collected
    chart = BarChart()
    chart.type    = "col"
    chart.title   = "GPR vs Collected Rent"
    chart.y_axis.title = "Revenue ($)"
    chart.x_axis.title = "Month"
    chart.style   = 10
    chart.width   = 22
    chart.height  = 12

    data_ref = Reference(ws, min_col=2, max_col=3,
                         min_row=5, max_row=5+len(df))
    cats_ref = Reference(ws, min_col=1,
                         min_row=6, max_row=5+len(df))
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = DARK_GREEN
    chart.series[1].graphicalProperties.solidFill = ACCENT_GOLD

    ws.add_chart(chart, "G5")
    set_col_widths(ws, {"A":12,"B":22,"C":18,"D":14,"E":16})


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — MAINTENANCE ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════
def sheet_maintenance(wb, conn):
    ws = wb.create_sheet("Maintenance Analysis")
    ws.sheet_view.showGridLines = False
    title_block(ws, "Maintenance Cost & SLA Analysis",
                "Open tickets, resolution times, cost by property & category")

    df = pd.read_sql("""
    SELECT
        p.name property, m.category, m.priority,
        COUNT(m.ticket_id) total_tickets,
        SUM(CASE WHEN m.status='Open' THEN 1 ELSE 0 END) open_tickets,
        ROUND(AVG(CASE WHEN m.close_date IS NOT NULL
              THEN julianday(m.close_date)-julianday(m.open_date) END),1) avg_resolution_days,
        ROUND(SUM(m.cost),0) total_cost,
        ROUND(AVG(m.cost),0) avg_cost
    FROM maintenance m
    JOIN properties p ON m.property_id = p.property_id
    GROUP BY p.property_id, m.category, m.priority
    ORDER BY total_cost DESC
    """, conn)

    cols = ["Property","Category","Priority","Total Tickets",
            "Open Tickets","Avg Resolution Days","Total Cost","Avg Cost"]
    header_row(ws, 5, cols)

    priority_colors = {
        "Emergency": (RED_LIGHT, RED_DARK),
        "High":      (AMBER_LIGHT, AMBER_DARK),
    }
    for i, (_, row) in enumerate(df.iterrows()):
        bg = LIGHT_GRAY if i % 2 == 0 else WHITE
        data_row(ws, 6+i,
                 [row["property"],row["category"],row["priority"],
                  row["total_tickets"],row["open_tickets"],
                  row["avg_resolution_days"],row["total_cost"],row["avg_cost"]],
                 bg=bg, fmt_map={7:"$#,##0",8:"$#,##0"})
        pri_cell = ws.cell(row=6+i, column=3)
        if row["priority"] in priority_colors:
            bg_c, fg_c = priority_colors[row["priority"]]
            pri_cell.fill = fill(bg_c)
            pri_cell.font = font(color=fg_c, bold=True)

    set_col_widths(ws, {"A":24,"B":14,"C":12,"D":13,
                        "E":12,"F":20,"G":12,"H":10})


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — FEATURE ENGINEERING INSIGHTS
# ══════════════════════════════════════════════════════════════════════════════
def sheet_feature_insights(wb):
    ws = wb.create_sheet("Feature Engineering")
    ws.sheet_view.showGridLines = False
    title_block(ws, "Feature-Engineered Metrics",
                "Derived analytics: risk scores, loss-to-lease, SLA breaches")

    try:
        df = pd.read_csv("data/feat_property_kpis.csv")
        cols = ["name","city","occupancy_rate","vacancy_rate",
                "avg_market_rent","avg_actual_rent","loss_to_lease_pct",
                "collection_rate_pct","revenue_per_unit",
                "performance_score","performance_grade"]
        df = df[[c for c in cols if c in df.columns]]

        header_row(ws, 5,
                   ["Property","City","Occupancy %","Vacancy %",
                    "Avg Market Rent","Avg Actual Rent","Loss-to-Lease %",
                    "Collection Rate %","Rev/Unit","Perf Score","Grade"])

        grade_colors = {
            "A+": (LIGHT_GREEN, MID_GREEN),
            "A":  (LIGHT_GREEN, MID_GREEN),
            "B":  (AMBER_LIGHT, AMBER_DARK),
            "C":  (RED_LIGHT,   RED_DARK),
        }
        for i, (_, row) in enumerate(df.iterrows()):
            bg = LIGHT_GRAY if i % 2 == 0 else WHITE
            vals = [row.get(c,"") for c in df.columns]
            data_row(ws, 6+i, vals, bg=bg,
                     fmt_map={3:"0.0%",4:"0.0%",5:"$#,##0",6:"$#,##0",
                               7:"0.00%",8:"0.0%",9:"$#,##0"})
            grade_cell = ws.cell(row=6+i, column=len(df.columns))
            g = str(row.get("performance_grade",""))
            if g in grade_colors:
                bg_c, fg_c = grade_colors[g]
                grade_cell.fill = fill(bg_c)
                grade_cell.font = font(color=fg_c, bold=True)

        set_col_widths(ws, {"A":24,"B":14,"C":12,"D":10,"E":16,
                            "F":16,"G":14,"H":16,"I":12,"J":12,"K":8})
    except FileNotFoundError:
        ws.cell(row=5, column=1,
                value="Run feature_engineering.py first to populate this sheet.")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN — BUILD WORKBOOK
# ══════════════════════════════════════════════════════════════════════════════
def main():
    print("\n" + "="*55)
    print("  Sage Ventures — Phase 4: Excel Report Generator")
    print("="*55)

    conn = sqlite3.connect(DB_PATH)
    wb   = Workbook()
    wb.remove(wb.active)  # remove default sheet

    print("  Building Sheet 1: Executive Summary...")
    sheet_executive_summary(wb, conn)

    print("  Building Sheet 2: Rent Roll...")
    sheet_rent_roll(wb, conn)

    print("  Building Sheet 3: Lease Expiration Pipeline...")
    sheet_lease_expiration(wb, conn)

    print("  Building Sheet 4: Delinquency Report...")
    sheet_delinquency(wb, conn)

    print("  Building Sheet 5: YSR Revenue Trend...")
    sheet_ysr(wb, conn)

    print("  Building Sheet 6: Maintenance Analysis...")
    sheet_maintenance(wb, conn)

    print("  Building Sheet 7: Feature Engineering Insights...")
    sheet_feature_insights(wb)

    conn.close()
    wb.save(OUT_FILE)

    print("="*55)
    print("  Phase 4 Complete ✓")
    print(f"  Excel Report → {OUT_FILE}")
    print("  7 sheets | Charts | Conditional formatting | KPI cards")
    print("="*55)

if __name__ == "__main__":
    main()
